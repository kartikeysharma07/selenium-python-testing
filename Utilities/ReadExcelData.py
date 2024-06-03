from configparser import ConfigParser

import openpyxl, os


class DataProvider:
    def __init__(self):
        config=ConfigParser()
        config.read(r"../Utilities/fileLocations.ini")
        self.excel_file_path = config.get('paths', 'excel_file_location')
        self.bk = openpyxl.load_workbook(self.excel_file_path)

    def get_login_credentials(self):
        credentials = []
        sheet=self.bk["login"]
        for m in range(2, sheet.max_row + 1):
            username = sheet.cell(row=m, column=1).value
            password = sheet.cell(row=m, column=2).value
            credentials.append((username, password))
        return credentials

    def get_register_credentials(self):
        credentials=[]
        sheet=self.bk["register"]
        for m in range(2,sheet.max_row+1):
            for n in range(1, sheet.max_column+1):
                row_data = {
                    'name': sheet.cell(row=m, column=1).value,
                    'email': sheet.cell(row=m, column=2).value,
                    'password': sheet.cell(row=m, column=3).value,
                    'firstname':sheet.cell(row=m,column=4).value,
                    'lastname':sheet.cell(row=m,column=5).value,
                    'address':sheet.cell(row=m,column=6).value,
                    'country':sheet.cell(row=m,column=7).value,
                    'state':sheet.cell(row=m,column=8).value,
                    'city':sheet.cell(row=m,column=9).value,
                    'zipcode':sheet.cell(row=m,column=10).value,
                    'mobile':sheet.cell(row=m,column=11).value,
                }
            credentials.append(row_data)
        return credentials


    def get_payment_details(self):
        details=[]
        sheet=self.bk["payment"]
        for m in range(2,sheet.max_row+1):
            for n in range(1,sheet.max_column+1):
                row_data={
                    'name':sheet.cell(row=m,column=1).value,
                    'card number':sheet.cell(row=m,column=2).value,
                    'cvc':sheet.cell(row=m,column=3).value,
                    'expiry-month':sheet.cell(row=m,column=4).value,
                    'expiry-year':sheet.cell(row=m,column=5).value
                }
            details.append(row_data)
        return details

    def get_products(self):
        details=[]
        sheet=self.bk["products"]
        for m in range(2,sheet.max_row+1):
            product=sheet.cell(row=m,column=1).value
            details.append(product)
        return details